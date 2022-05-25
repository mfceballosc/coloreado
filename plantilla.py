"""
La información fue extraida del siguiente enlace
https://programmerclick.com/article/2492972998/
"""



from turtle import left, right
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, colors, Font, Alignment, PatternFill


if __name__=='__main__':
    #****************** carga de la tabla de excel ******************
    wb = load_workbook('Libro1.xlsx')

    #****************** Selecciona la hoja **************************
    ws: object = wb['sheet1']

    #****************** Establecer contenido de la celda ************
    ws.cell(1,1).value = 'prueba 1'
    ws.cell(1,2).value = 'prueba 2'

    #****************** Establecer color del fondo de la celda ******
    color_celda = PatternFill(fill_type='solid', fgColor="D3D3D3")
    ws.cell(1,1).fill = color_celda

    #****************** Combinar celdas *****************************
    ws.merge_cells('f1:h3')

    #****************** Colorear celda combinada ********************
    ws['f4'].fill = color_celda
    ws.merge_cells('f4:h7')

    #****************** Establecer el bordes ************************
    # define los bordes
    _left = Side(border_style='thin', color=colors.BLACK)
    _right = Side(border_style='thin', color=colors.BLACK)
    _top = Side(border_style='thin', color=colors.BLACK)
    _bottom = Side(border_style='thin', color=colors.BLACK)

    # Borde de una celda
    border = Border(left=_left, right=_right, top=_top, bottom=_bottom)
    ws["i1"].border = border
    # Borde de una celda combinada
    ws["i2"].border = border
    ws.merge_cells('i2:i7')

    #****************** fuente y estilo de alineación ***************
    font_content = Font('Arial', 
                        size=11, 
                        color=colors.BLACK, 
                        bold=True, 
                        italic=True)

    alignment_cc = Alignment(horizontal='center', 
                            vertical='center', 
                            text_rotation=0, 
                            wrap_text=True, 
                            shrink_to_fit=True, 
                            indent=0)
    alignment_cl = Alignment(horizontal='left', 
                            vertical='center', 
                            text_rotation=0, 
                            wrap_text=True, 
                            shrink_to_fit=True, 
                            indent=0)
    ws.cell(1,1).font = font_content
    ws.cell(1,1).alignment=alignment_cc
    ws.cell(1,2).font = font_content
    ws.cell(1,2).alignment=alignment_cl

    print(ws)


    wb.save('Nuevo Form.xlsx')