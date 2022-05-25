


from fileinput import filename
import openpyxl as opxl
from openpyxl.styles import PatternFill as ptfill
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl import load_workbook, Workbook




if __name__=='__main__':
    tabla = opxl.load_workbook('Libro1.xlsx', data_only=True)
    hoja1 = tabla.active
    colores = COLOR_INDEX
    idx_color = 63
    hoja1["B3"].fill = ptfill(bgColor=colores[idx_color],  fill_type="solid")
    hoja1["B3"].font = rojo20

    tabla.save(filename='Libro1.xlsx')

    print(len(colores), colores)

